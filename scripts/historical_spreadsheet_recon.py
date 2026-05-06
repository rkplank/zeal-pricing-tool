"""Read-only reconnaissance for historical Zeal gift-card pricing workbooks.

The output is analysis material only: JSON row artifacts plus a Markdown summary.
It does not write to SQLite or import the live pricing engine.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from collections import Counter, defaultdict
from collections.abc import Iterable
from dataclasses import asdict, dataclass
from datetime import UTC, datetime
from pathlib import Path
from typing import Any, Protocol

import xlrd
from openpyxl import load_workbook

try:
    from scripts.spreadsheet_parser import slugify
except ModuleNotFoundError:  # pragma: no cover - direct script execution path
    from spreadsheet_parser import slugify

DEFAULT_WORKBOOKS = [
    Path(r"C:\Users\kandt\Downloads\GiftCardPricingData_2017.xls"),
    Path(r"C:\Users\kandt\Downloads\GiftCardPricingData_2018.xls"),
    Path(r"C:\Users\kandt\Downloads\GiftCardPricingData_2021.xls"),
    Path(r"C:\Users\kandt\Downloads\GiftCardPricingData_2022.xls"),
    Path(r"C:\Users\kandt\Downloads\GiftCardPricingData_2023.xls"),
    Path(r"C:\Users\kandt\Downloads\GiftCardPricingData_2025.xls"),
    Path(r"C:\Users\kandt\OneDrive - Umich\Zeal Cards\Pricing\GiftCardPricingData_2025.xlsx"),
]

DEFAULT_ROWS_JSON = Path("docs/generated/historical_pricing_rows.json")
DEFAULT_INVENTORY_JSON = Path("docs/generated/historical_workbook_inventory.json")
DEFAULT_REPORT = Path("docs/historical_pricing_analysis.md")

CANONICAL_COLUMNS = (
    "merchant_name",
    "ebay_sell",
    "online_sell",
    "in_mail_buy",
    "in_store_buy",
    "electronic_buy",
    "tier",
    "cardcash_buy",
    "cardcash_sell",
    "raise_sell",
)

KNOWN_TIERS = {"T24", "C", "Z", "NC"}
SECTION_NAMES = {
    "merchant",
    "physical only",
    "bankrupt",
    "bread and butter",
    "top cards",
    "covid watch list",
    "take online",
    "local",
}


@dataclass(frozen=True)
class CellValue:
    value: Any
    formula: str | None = None


class SheetReader(Protocol):
    name: str
    nrows: int
    ncols: int

    def cell(self, row: int, col: int) -> CellValue: ...


@dataclass(frozen=True)
class UsedRange:
    first_row: int | None
    last_row: int | None
    first_col: int | None
    last_col: int | None

    @property
    def display(self) -> str:
        if self.first_row is None:
            return "(empty)"
        return (
            f"R{self.first_row + 1}C{self.first_col + 1}:"
            f"R{self.last_row + 1}C{self.last_col + 1}"
        )


@dataclass(frozen=True)
class SheetRecon:
    name: str
    used_range: UsedRange
    likely_pricing_sheet: bool
    header_row: int | None
    detected_columns: dict[str, int]
    data_row_count: int


@dataclass(frozen=True)
class NormalizedRow:
    year: int
    workbook: str
    workbook_format: str
    sheet: str
    source_row: int
    merchant_name: str
    merchant_id: str
    ebay_sell: float | None = None
    online_sell: float | None = None
    in_mail_buy: float | None = None
    in_store_buy: float | None = None
    electronic_buy: float | None = None
    tier: str | None = None
    cardcash_buy: float | None = None
    cardcash_sell: float | None = None
    raise_sell: float | None = None
    notes: tuple[str, ...] = ()


def normalize_label(value: object) -> str:
    text = "" if value is None else str(value)
    text = text.strip().lower()
    text = re.sub(r"[\r\n\t]+", " ", text)
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def parse_percent(value: object) -> float | None:
    """Normalize spreadsheet percentage-like values into fractions in [0, 1.2]."""
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        numeric = float(value)
    elif isinstance(value, str):
        text = value.strip()
        if not text or text.lower() in {"no", "n/a", "na", "no data", "#value!"}:
            return None
        text = text.replace("%", "").strip()
        try:
            numeric = float(text)
        except ValueError:
            return None
        if "%" in value or numeric > 1.2:
            numeric /= 100.0
    else:
        return None
    if 0 <= numeric <= 1.2:
        return numeric
    return None


def extract_year(path: Path) -> int:
    match = re.search(r"(20\d{2})", path.name)
    if not match:
        raise ValueError(f"could not extract year from {path.name}")
    return int(match.group(1))


def detect_column_map(header_values: list[object]) -> dict[str, int]:
    """Return canonical column names mapped to zero-based column indexes."""
    normalized = [normalize_label(value) for value in header_values]
    result: dict[str, int] = {}

    for idx, label in enumerate(normalized):
        if not label:
            continue
        if "merchant" in label and "merchant_name" not in result:
            result["merchant_name"] = idx
        elif "ebay" in label and "sell" in label:
            result.setdefault("ebay_sell", idx)
        elif "online" in label and ("sell" in label or "buy" in label):
            result.setdefault("online_sell", idx)
        elif "in mail" in label or "mail" == label:
            result.setdefault("in_mail_buy", idx)
        elif "in store" in label or "store" == label:
            result.setdefault("in_store_buy", idx)
        elif "electronic" in label or label in {"e", "e buy"}:
            result.setdefault("electronic_buy", idx)
        elif label == "tier" or "tier" in label:
            result.setdefault("tier", idx)
        elif "cardcash" in label and "sell" in label:
            result.setdefault("cardcash_sell", idx)
        elif "cardcash" in label and ("buy" in label or "cash" in label):
            result.setdefault("cardcash_buy", idx)
        elif "raise" in label and "sell" in label:
            result.setdefault("raise_sell", idx)

    return result


def pricing_sheet_column_map(sheet: SheetReader) -> dict[str, int]:
    """Known split-header layout for historical PricingSheet tabs."""
    if sheet.ncols < 7:
        return {}
    row_one = [normalize_label(sheet.cell(0, col).value) for col in range(sheet.ncols)]
    row_two = [normalize_label(sheet.cell(1, col).value) for col in range(sheet.ncols)]
    if row_two[0] != "merchant" or "in mail" not in row_two[2]:
        return {}
    result = {
        "merchant_name": 0,
        "ebay_sell": 1,
        "in_mail_buy": 2,
        "electronic_buy": 3,
        "in_store_buy": 4,
        "online_sell": 5,
        "tier": 6,
    }
    for idx, label in enumerate(row_two):
        if "cc sell" in label or "cardcash sell" in label:
            result["cardcash_sell"] = idx
        if "cc buy" in label or "cardcash buy" in label:
            result["cardcash_buy"] = idx
        if "raise" in label and "sell" in label:
            result["raise_sell"] = idx
    for idx, label in enumerate(row_one):
        if "raise" in label and "sell" in label:
            result["raise_sell"] = idx
    return result


def is_probable_merchant_name(value: object) -> bool:
    if not isinstance(value, str):
        return False
    text = value.strip()
    if not text:
        return False
    normalized = normalize_label(text)
    if normalized in SECTION_NAMES:
        return False
    if normalized in {"n a", "no data", "no"}:
        return False
    return any(ch.isalpha() for ch in text)


def used_range(sheet: SheetReader) -> UsedRange:
    first_row: int | None = None
    last_row: int | None = None
    first_col: int | None = None
    last_col: int | None = None
    for row in range(sheet.nrows):
        for col in range(sheet.ncols):
            cell = sheet.cell(row, col)
            if cell.value is None or cell.value == "":
                continue
            first_row = row if first_row is None else min(first_row, row)
            last_row = row if last_row is None else max(last_row, row)
            first_col = col if first_col is None else min(first_col, col)
            last_col = col if last_col is None else max(last_col, col)
    return UsedRange(first_row, last_row, first_col, last_col)


def find_header_row(sheet: SheetReader) -> tuple[int | None, dict[str, int]]:
    if normalize_label(sheet.name) == "pricingsheet":
        pricing_map = pricing_sheet_column_map(sheet)
        if pricing_map:
            return 1, pricing_map

    best_row: int | None = None
    best_map: dict[str, int] = {}
    for row in range(min(sheet.nrows, 25)):
        values = [sheet.cell(row, col).value for col in range(sheet.ncols)]
        column_map = detect_column_map(values)
        score = len(set(column_map) & {"merchant_name", "ebay_sell", "in_mail_buy", "tier"})
        if score > len(set(best_map) & {"merchant_name", "ebay_sell", "in_mail_buy", "tier"}):
            best_row = row
            best_map = column_map
    if "merchant_name" not in best_map:
        return None, {}
    return best_row, best_map


def extract_rows(
    sheet: SheetReader,
    workbook_path: Path,
    workbook_format: str,
    header_row: int,
    column_map: dict[str, int],
) -> list[NormalizedRow]:
    year = extract_year(workbook_path)
    rows: list[NormalizedRow] = []
    seen_ids: Counter[str] = Counter()

    merchant_col = column_map["merchant_name"]
    for row_idx in range(header_row + 1, sheet.nrows):
        merchant_value = sheet.cell(row_idx, merchant_col).value
        if not is_probable_merchant_name(merchant_value):
            continue
        merchant_name = str(merchant_value).strip()
        base_id = slugify(merchant_name)
        seen_ids[base_id] += 1
        merchant_id = base_id if seen_ids[base_id] == 1 else f"{base_id}_{seen_ids[base_id]}"
        notes: list[str] = []
        if seen_ids[base_id] > 1:
            notes.append(f"duplicate merchant slug {base_id}")

        def raw(canonical: str) -> object:
            col = column_map.get(canonical)
            return sheet.cell(row_idx, col).value if col is not None and col < sheet.ncols else None

        tier_raw = raw("tier")
        tier = str(tier_raw).strip() if str(tier_raw).strip() in KNOWN_TIERS else None
        if tier is None:
            continue

        rows.append(
            NormalizedRow(
                year=year,
                workbook=workbook_path.name,
                workbook_format=workbook_format,
                sheet=sheet.name,
                source_row=row_idx + 1,
                merchant_name=merchant_name,
                merchant_id=merchant_id,
                ebay_sell=parse_percent(raw("ebay_sell")),
                online_sell=parse_percent(raw("online_sell")),
                in_mail_buy=parse_percent(raw("in_mail_buy")),
                in_store_buy=parse_percent(raw("in_store_buy")),
                electronic_buy=parse_percent(raw("electronic_buy")),
                tier=tier,
                cardcash_buy=parse_percent(raw("cardcash_buy")),
                cardcash_sell=parse_percent(raw("cardcash_sell")),
                raise_sell=parse_percent(raw("raise_sell")),
                notes=tuple(notes),
            )
        )
    return rows


class XlrdSheet:
    def __init__(self, sheet: xlrd.sheet.Sheet) -> None:
        self._sheet = sheet
        self.name = sheet.name
        self.nrows = sheet.nrows
        self.ncols = sheet.ncols

    def cell(self, row: int, col: int) -> CellValue:
        if row >= self.nrows or col >= self.ncols:
            return CellValue(None)
        cell = self._sheet.cell(row, col)
        if cell.ctype in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK):
            return CellValue(None)
        if cell.ctype == xlrd.XL_CELL_ERROR:
            return CellValue(xlrd.error_text_from_code.get(cell.value, "#ERROR"))
        return CellValue(cell.value)


class OpenpyxlSheet:
    def __init__(self, values_sheet: Any, formulas_sheet: Any) -> None:
        self._values_sheet = values_sheet
        self._formulas_sheet = formulas_sheet
        self.name = values_sheet.title
        self.nrows = values_sheet.max_row or 0
        self.ncols = values_sheet.max_column or 0

    def cell(self, row: int, col: int) -> CellValue:
        values_cell = self._values_sheet.cell(row=row + 1, column=col + 1)
        formulas_cell = self._formulas_sheet.cell(row=row + 1, column=col + 1)
        formula = formulas_cell.value if isinstance(formulas_cell.value, str) else None
        formula_value = formula if formula and formula.startswith("=") else None
        return CellValue(values_cell.value, formula_value)


def iter_sheets(path: Path) -> tuple[str, list[SheetReader]]:
    suffix = path.suffix.lower()
    if suffix == ".xls":
        workbook = xlrd.open_workbook(path, on_demand=True)
        return "xls", [XlrdSheet(workbook.sheet_by_name(name)) for name in workbook.sheet_names()]
    if suffix == ".xlsx":
        wb_values = load_workbook(path, data_only=True, read_only=False)
        wb_formulas = load_workbook(path, data_only=False, read_only=False)
        return "xlsx", [
            OpenpyxlSheet(wb_values[name], wb_formulas[name]) for name in wb_values.sheetnames
        ]
    raise ValueError(f"unsupported workbook format: {path}")


def inspect_workbook(path: Path) -> tuple[list[SheetRecon], list[NormalizedRow]]:
    workbook_format, sheets = iter_sheets(path)
    sheet_recons: list[SheetRecon] = []
    normalized_rows: list[NormalizedRow] = []

    for sheet in sheets:
        sheet_range = used_range(sheet)
        header_row, column_map = find_header_row(sheet)
        data_rows: list[NormalizedRow] = []
        if header_row is not None and "merchant_name" in column_map:
            data_rows = extract_rows(sheet, path, workbook_format, header_row, column_map)
        likely = normalize_label(sheet.name) == "pricingsheet" and len(data_rows) >= 20
        sheet_recons.append(
            SheetRecon(
                name=sheet.name,
                used_range=sheet_range,
                likely_pricing_sheet=likely,
                header_row=header_row + 1 if header_row is not None else None,
                detected_columns={key: value + 1 for key, value in sorted(column_map.items())},
                data_row_count=len(data_rows),
            )
        )
        if likely:
            normalized_rows.extend(data_rows)
    return sheet_recons, normalized_rows


def available_columns(rows: Iterable[NormalizedRow]) -> dict[str, bool]:
    rows_list = list(rows)
    return {
        column: any(getattr(row, column) is not None for row in rows_list)
        for column in CANONICAL_COLUMNS
        if column != "merchant_name"
    }


def detected_columns_for_year(
    year: int,
    workbook_recons: dict[str, list[SheetRecon]],
) -> dict[str, bool]:
    detected = set()
    for workbook_name, sheet_recons in workbook_recons.items():
        if extract_year(Path(workbook_name)) != year:
            continue
        for sheet in sheet_recons:
            detected.update(sheet.detected_columns)
    return {
        column: column in detected
        for column in CANONICAL_COLUMNS
        if column != "merchant_name"
    }


def build_report(
    workbook_recons: dict[str, list[SheetRecon]],
    rows: list[NormalizedRow],
    generated_at: str,
) -> str:
    rows_by_year: dict[int, list[NormalizedRow]] = defaultdict(list)
    for row in rows:
        rows_by_year[row.year].append(row)

    merchant_sets = {
        year: {row.merchant_id for row in year_rows}
        for year, year_rows in rows_by_year.items()
    }
    years = sorted(merchant_sets)
    common_all = set.intersection(*(merchant_sets[year] for year in years)) if years else set()

    display_by_id: dict[str, str] = {}
    for row in rows:
        display_by_id.setdefault(row.merchant_id, row.merchant_name)

    lines = [
        "# Historical Pricing Workbook Analysis",
        "",
        f"Generated: {generated_at}",
        "",
        "This is read-only reconnaissance. The extracted JSON artifacts live under "
        "`docs/generated/` and are not production database tables.",
        "",
        "## Workbook Inventory",
        "",
        "| Workbook | Format | Sheets | Likely pricing sheets | Used ranges |",
        "|---|---:|---|---|---|",
    ]

    for workbook_name, sheet_recons in workbook_recons.items():
        formats = workbook_name.rsplit(".", 1)[-1]
        likely = [sheet.name for sheet in sheet_recons if sheet.likely_pricing_sheet]
        ranges = "; ".join(f"{sheet.name}: {sheet.used_range.display}" for sheet in sheet_recons)
        lines.append(
            f"| {workbook_name} | {formats} | "
            f"{', '.join(sheet.name for sheet in sheet_recons)} | "
            f"{', '.join(likely) or '(none)'} | {ranges} |"
        )

    lines += [
        "",
        "## Merchant Counts By Year",
        "",
        "| Year | Merchant rows | Unique normalized merchants | Likely pricing sheets |",
        "|---:|---:|---:|---|",
    ]
    for year in years:
        year_rows = rows_by_year[year]
        sheets = sorted({row.sheet for row in year_rows})
        lines.append(
            f"| {year} | {len(year_rows)} | "
            f"{len(merchant_sets[year])} | {', '.join(sheets)} |"
        )

    lines += [
        "",
        "## Columns Available By Year",
        "",
        "| Year | eBay sell | Online sell | In-mail buy | In-store buy | "
        "Electronic buy | Tier | CardCash buy | CardCash sell | Raise sell |",
        "|---:|---|---|---|---|---|---|---|---|---|",
    ]
    for year in years:
        cols = detected_columns_for_year(year, workbook_recons)

        def marker(name: str) -> str:
            return "yes" if cols.get(name) else "no"

        lines.append(
            f"| {year} | {marker('ebay_sell')} | {marker('online_sell')} | "
            f"{marker('in_mail_buy')} | {marker('in_store_buy')} | "
            f"{marker('electronic_buy')} | {marker('tier')} | "
            f"{marker('cardcash_buy')} | {marker('cardcash_sell')} | {marker('raise_sell')} |"
        )

    lines += [
        "",
        "## Common Merchants Across Years",
        "",
        f"{len(common_all)} normalized merchants appear in every inspected year.",
    ]
    if common_all:
        common_names = sorted(display_by_id[mid] for mid in common_all)
        preview = ", ".join(common_names[:80])
        suffix = "" if len(common_names) <= 80 else f" ... ({len(common_names) - 80} more)"
        lines.append("")
        lines.append(preview + suffix)

    lines += ["", "## Merchants Added/Removed", ""]
    for prev, curr in zip(years, years[1:], strict=False):
        added = merchant_sets[curr] - merchant_sets[prev]
        removed = merchant_sets[prev] - merchant_sets[curr]
        lines.append(f"### {prev} to {curr}")
        lines.append("")
        lines.append(f"Added: {len(added)}")
        lines.append(_format_name_list(added, display_by_id))
        lines.append("")
        lines.append(f"Removed: {len(removed)}")
        lines.append(_format_name_list(removed, display_by_id))
        lines.append("")

    lines += [
        "## Obvious Schema Differences",
        "",
        "- The historical `.xls` workbooks expose cached values but not formulas through "
        "the reader used here, so formula-reference drift cannot be reconstructed from "
        "those files without an Excel/LibreOffice conversion step.",
        "- The 2025 `.xlsx` workbook exposes formulas and includes the known "
        "`PricingSheet` and `InputsandMargins` structure used by the live v1 "
        "baseline parser.",
    ]
    for year in years:
        cols = detected_columns_for_year(year, workbook_recons)
        missing = [name for name, present in cols.items() if not present]
        if missing:
            lines.append(f"- {year}: no detected values for {', '.join(missing)}.")

    duplicate_counts = Counter((row.year, row.merchant_id) for row in rows)
    duplicates = [(year, mid) for (year, mid), count in duplicate_counts.items() if count > 1]
    if duplicates:
        lines.append(
            "- Duplicate normalized merchant slugs exist within at least one workbook; "
            "see row artifact notes."
        )

    lines += [
        "",
        "## Risks/Unknowns",
        "",
        "- Column detection is heuristic for historical sheets; values should be "
        "spot-checked before using this for any migration or trend analysis.",
        "- Merchant matching is slug-based. Renames, punctuation changes, and grouped "
        "merchants can appear as added/removed even when the business entity is "
        "continuous.",
        "- `.xls` formula text is unavailable with the current lightweight reader, so "
        "margin-row, differential-row, and e-bonus provenance is unknown for "
        "2017-2025 `.xls` files.",
        "- Extracted percentages are normalized as fractions, but unusual text markers "
        "and workbook errors are preserved only as missing values plus notes.",
        "",
        "## Recommended Next Analysis Steps",
        "",
        "1. Spot-check 10-15 high-volume merchants across all years against the "
        "original workbooks.",
        "2. Build a manual rename map for merchant names that changed but represent "
        "the same merchant.",
        "3. Convert the `.xls` workbooks to `.xlsx` on a controlled machine if "
        "formula-reference history matters.",
        "4. Compare historical channel percentages for stable common merchants to "
        "identify broad policy shifts, without changing the live v1 algorithm.",
    ]
    return "\n".join(lines) + "\n"


def _format_name_list(merchant_ids: set[str], display_by_id: dict[str, str]) -> str:
    if not merchant_ids:
        return "(none)"
    names = [display_by_id.get(mid, mid) for mid in sorted(merchant_ids)]
    if len(names) > 40:
        return ", ".join(names[:40]) + f" ... ({len(names) - 40} more)"
    return ", ".join(names)


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("workbooks", nargs="*", type=Path, default=DEFAULT_WORKBOOKS)
    parser.add_argument("--rows-json", type=Path, default=DEFAULT_ROWS_JSON)
    parser.add_argument("--inventory-json", type=Path, default=DEFAULT_INVENTORY_JSON)
    parser.add_argument("--report", type=Path, default=DEFAULT_REPORT)
    args = parser.parse_args()

    generated_at = datetime.now(UTC).strftime("%Y-%m-%dT%H:%M:%SZ")
    all_rows: list[NormalizedRow] = []
    workbook_recons: dict[str, list[SheetRecon]] = {}

    for workbook in args.workbooks:
        if not workbook.exists():
            print(f"warning: workbook not found, skipping: {workbook}", file=sys.stderr)
            continue
        sheet_recons, rows = inspect_workbook(workbook)
        workbook_recons[workbook.name] = sheet_recons
        all_rows.extend(rows)

    args.rows_json.parent.mkdir(parents=True, exist_ok=True)
    args.inventory_json.parent.mkdir(parents=True, exist_ok=True)
    args.report.parent.mkdir(parents=True, exist_ok=True)

    args.rows_json.write_text(
        json.dumps([asdict(row) for row in all_rows], indent=2, sort_keys=True),
        encoding="utf-8",
    )
    args.inventory_json.write_text(
        json.dumps(
            {
                name: [
                    {
                        **asdict(sheet),
                        "used_range": asdict(sheet.used_range)
                        | {"display": sheet.used_range.display},
                    }
                    for sheet in sheets
                ]
                for name, sheets in workbook_recons.items()
            },
            indent=2,
            sort_keys=True,
        ),
        encoding="utf-8",
    )
    args.report.write_text(build_report(workbook_recons, all_rows, generated_at), encoding="utf-8")

    print(f"wrote {args.rows_json} ({len(all_rows)} rows)")
    print(f"wrote {args.inventory_json}")
    print(f"wrote {args.report}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
